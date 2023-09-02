import os
import re
from typing import List, Tuple

import openpyxl
from tqdm import tqdm

from init import WORKDIR, data_clean, data_win_file, get_custom_filename_with_params

class QQtoExcel:
    """
    将QQ聊天记录导出为Excel文件的类。
    """

    def __init__(self, qq_chat_route: str = None, file_path: str = None,
                 sheet_name: str = "[消息对象]", rule_string: str = '',
                 time_list_out: bool = True, name_list_out: bool = True,
                 uid_list_out: bool = True, cont_list_out: bool = True,
                 cont_nil_out: bool = False, multi_sheet_export: int = 0,
                 time_row_text: str = "时间", name_row_text: str = "昵称",
                 uid_row_text: str = "QQ（邮箱）", cont_row_text: str = "内容",
                 out_type: int = 0) -> None:
        """
        初始化 QQtoExcel 类。

        :param qq_chat_route: QQ聊天记录路径
        :param file_path: 导出目录
        :param sheet_name: 工作表的名字
        :param rule_string: 自定义文件名规则
        :param time_list_out: 是否导出时间
        :param name_list_out: 是否导出昵称
        :param uid_list_out: 是否导出Uid
        :param cont_list_out: 是否导出内容
        :param cont_nil_out: 是否过滤无意义内容
        :param multi_sheet_export: 多工作表导出模式（0：非多工作表导出, 1：按消息对象导出, 2：按分组导出）
        :param time_row_text: "时间"列表头的标题
        :param name_row_text: "昵称"列表头的标题
        :param uid_row_text: "QQ（邮箱）"列表头的标题
        :param cont_row_text: "内容"列表头的标题
        :param out_type: 导出模式，0 为按好友导出，非0 按分组导出
        """
        self.qq_chat_route = qq_chat_route or os.path.join(WORKDIR, '全部消息记录.txt')
        self.file_path = file_path or os.path.join(WORKDIR, "out")
        self.sheet_name = sheet_name
        self.rule_string = rule_string.strip()
        self.time_row_text = time_row_text
        self.name_row_text = name_row_text
        self.uid_row_text = uid_row_text
        self.cont_row_text = cont_row_text
        self.cont_nil_out = cont_nil_out
        self.multi_sheet_export = multi_sheet_export
        self.out_type = out_type
        self.row = []

        if self.out_type != 0 and self.multi_sheet_export == 1:
            raise ValueError(
                "导出模式与多工作表导出模式冲突，您不能在选择按消息分组导出的同时选择多工作表按消息对象导出。")

        if time_list_out:
            self.row.append(time_row_text)
        if name_list_out:
            self.row.append(name_row_text)
        if uid_list_out:
            self.row.append(uid_row_text)
        if cont_list_out:
            self.row.append(cont_row_text)

    def get_QQChat_record(self) -> Tuple[List[str], List[List[List[str]]]]:
        """
        读取 QQ 聊天记录并生成对象。

        :return: 生成的对象列表和文件名列表
        :rtype: tuple[list[str], list[list[list[str]]]]
        """
        object_file_name_list = []  # 文件名称列表（消息分组_消息对象 or 消息分组/消息对象）
        object_list = []  # 消息对象列表

        time_list = []
        name_list = []
        uid_list = []
        cont_list = []

        with open(self.qq_chat_route, encoding="utf-8") as f:
            text = f.read()

            pattern = re.compile(r"^\n^={64}\n^消息分组:(.+?)\n={64}\n^消息对象:(.+?)\n^={64}$",
                                 re.DOTALL | re.MULTILINE)
            matches = pattern.finditer(text)

            matches_list = list(matches)  # 将迭代器转换为列表

            pattern_data = {}
            # 使用正则表达式来匹配消息分组与消息对象，并去重合并
            for i, match in enumerate(matches_list):
                start_pos = match.end()
                end_pos = matches_list[i + 1].start() if i + 1 < len(matches_list) else len(text)
                group = match[1]
                obj = match[2]
                key = (obj, group)  # 使用元组（obj，group）作为键

                if key in pattern_data:
                    pattern_data[key][1].append(text[start_pos:end_pos])
                else:
                    pattern_data[key] = (group, [text[start_pos:end_pos]])

            # 使用正则表达式来匹配每条聊天记录的时间、昵称、UID、内容
            message_pattern = re.compile(
                r'(\d{4}-\d{1,2}-\d{1,2}\s\d{1,2}:\d{1,2}:\d{1,2})(.+?([(|<](.*)[)|>])?\n)([\s\S]*?)(?=\n\s*(\d{4}-\d{1,2}-\d{1,2}\s\d{1,2}:\d{1,2}:\d{1,2})|$)')

            for (obj, group), (group_value, contents) in pattern_data.items():
                if len(obj) < 42:
                    if self.out_type == 0:
                        object_file_name_list.append(
                            data_win_file(data_clean(group) + "_" + data_clean(obj)))
                    else:
                        out_z_path_name = data_win_file(data_clean(group))
                        out_z_path = os.path.join(self.file_path, out_z_path_name)
                        if not os.path.exists(out_z_path):
                            os.mkdir(out_z_path)
                        object_file_name_list.append(out_z_path_name + "\\" + data_win_file(data_clean(obj)))

                    match_text_list = message_pattern.findall("\n".join(contents))
                    if match_text_list:
                        for j in match_text_list:
                            if not self.cont_row_text in self.row and self.cont_nil_out:
                                raise ValueError("无意义内容过滤选项与不导出内容选项冲突")
                            if self.cont_row_text in self.row:
                                cont_text = data_clean(j[4])
                                if self.cont_nil_out:
                                    cont_text = re.sub(f'\[(图片|语音|表情|QQ红包)]', "", cont_text)
                                    if len(cont_text.strip()) > 0:
                                        cont_list.append(cont_text)
                                    else:
                                        continue
                                else:
                                    cont_list.append(cont_text)
                            if self.time_row_text in self.row:
                                time_list.append(j[0])
                            if self.name_row_text in self.row:
                                name_list.append(
                                    data_clean(j[1].replace(j[2], '').replace('\\n', '')))
                            if self.uid_row_text in self.row:
                                uid_list.append(j[3])

                    object_list.append([time_list, name_list, uid_list, cont_list])
                    # 清空列表
                    time_list = []
                    name_list = []
                    uid_list = []
                    cont_list = []

        return object_file_name_list, object_list

    def _write_data_to_worksheet(self, worksheet, row_text, data_list):
        """
        将数据写入工作表。

        :param worksheet: 要写入的工作表对象
        :param row_text: 要写入的行标题
        :param data_list: 要写入的数据列表
        """
        if row_text in self.row:
            for k, data in enumerate(data_list):
                worksheet.cell(row=k + 2, column=self.row.index(row_text) + 1, value=data)

    def _create_excel_file(self, object_file_name_list, object_list):
        """
        创建Excel文件并导出数据。

        :param object_file_name_list: 文件名列表
        :param object_list: 消息对象列表
        """
        if len(self.row) <= 0:
            raise ValueError("请选择至少一项输出内容")

        files_path = {}  # 输出目录列表
        for i, object_file_name in enumerate(object_file_name_list):
            if self.out_type == 0:
                parts = object_file_name.split("_", 1)
            else:
                parts = object_file_name.split("\\", 1)
            list_name = parts[0]
            object_name = parts[1]
            if list_name in files_path:
                files_path[list_name].append((object_name, i))
            else:
                files_path[list_name] = [(object_name, i)]

        # 多工作表模式：按消息对象导出
        if self.multi_sheet_export == 1:
            # 设置默认自定义文件名规则
            if len(self.rule_string.strip()) <= 0:
                self.rule_string = 'output_[序号]'
            if len(object_file_name_list) > 255 and '[序号]' not in self.rule_string:
                raise ValueError("消息对象数量超过255个，您必须在自定义文件名规则中包含[序号]以防止导出文件名重复")
            if '[消息对象]' in self.rule_string or '[消息分组]' in self.rule_string:
                raise ValueError(
                    "在按消息对象导出的多工作表模式中，由于单个 Excel 中存在多个消息对象与消息分组，您无法使用文件命名规则中的[消息对象]与[消息分组]")

            current_workbook = openpyxl.Workbook()
            current_sheet_count = 0
            current_count = 0
            for list_name, object_name_list in files_path.items():
                for i in tqdm(range(len(object_name_list)), f'{list_name} 消息分组导出中'):
                    time_list, name_list, uid_list, cont_list = object_list[object_name_list[i][1]]

                    if current_sheet_count >= 255:
                        # 删除openpyxl创建的默认工作表
                        default_sheet = current_workbook['Sheet']
                        current_workbook.remove(default_sheet)
                        current_workbook.save(os.path.join(self.file_path, "".join(
                            [get_custom_filename_with_params(self.rule_string, list_name, '', current_count),
                             f'.xlsx'])))
                        current_workbook = openpyxl.Workbook()
                        current_count += 1
                        current_sheet_count = 0

                    worksheet = current_workbook.create_sheet(
                        title=get_custom_filename_with_params(self.sheet_name, list_name,
                                                              object_name_list[i][0], current_sheet_count))
                    current_sheet_count += 1

                    # 写入表头
                    for j, header in enumerate(self.row):
                        worksheet.cell(row=1, column=j + 1, value=header)

                    # 写入内容
                    # 使用新方法写入内容
                    self._write_data_to_worksheet(worksheet, self.time_row_text, time_list)
                    self._write_data_to_worksheet(worksheet, self.name_row_text, name_list)
                    self._write_data_to_worksheet(worksheet, self.uid_row_text, uid_list)
                    self._write_data_to_worksheet(worksheet, self.cont_row_text, cont_list)
            if current_sheet_count > 0:
                # 删除openpyxl创建的默认工作表
                default_sheet = current_workbook['Sheet']
                current_workbook.remove(default_sheet)
                current_workbook.save(os.path.join(self.file_path, "".join(
                    [get_custom_filename_with_params(self.rule_string, '', '', current_count), f'.xlsx'])))
        # 多工作表模式：按消息分组导出
        elif self.multi_sheet_export == 2:
            # 设置默认自定义文件名规则
            if len(self.rule_string.strip()) <= 0:
                self.rule_string = '[消息分组]_[序号]'
            if '[消息对象]' in self.rule_string:
                raise ValueError(
                    "在按消息分组导出的多工作表模式中，由于单个 Excel 中存在多个消息对象，您无法使用文件命名规则中的[消息对象]")

            for list_name, object_name_list in files_path.items():
                if len(object_name_list) > 255 and '[序号]' not in self.rule_string:
                    raise ValueError(
                        "存在消息分组中消息对象数量超过255个，您必须在自定义文件名规则中包含[序号]以防止导出文件名重复")
                current_workbook = openpyxl.Workbook()
                current_sheet_count = 0
                current_count = 0
                for i in tqdm(range(len(object_name_list)), f'{list_name} 消息分组导出中'):
                    time_list, name_list, uid_list, cont_list = object_list[object_name_list[i][1]]

                    if current_sheet_count >= 255:
                        # 删除openpyxl创建的默认工作表
                        default_sheet = current_workbook['Sheet']
                        current_workbook.remove(default_sheet)
                        # 保存当前工作簿并启动新工作簿
                        if self.out_type == 0:
                            current_workbook.save(os.path.join(self.file_path, "".join(
                                [get_custom_filename_with_params(self.rule_string, list_name, '', i // 255),
                                 f'.xlsx'])))
                        else:
                            current_workbook.save(os.path.join(self.file_path, "".join(
                                [list_name, '/',
                                 get_custom_filename_with_params(self.rule_string, list_name, '', i // 255),
                                 f'.xlsx'])))

                        current_workbook = openpyxl.Workbook()
                        current_sheet_count = 0
                        current_count += 1

                    worksheet = current_workbook.create_sheet(
                        title=get_custom_filename_with_params(self.sheet_name, list_name,
                                                              object_name_list[i][0], current_count))
                    current_sheet_count += 1

                    # 写入表头
                    for j, header in enumerate(self.row):
                        worksheet.cell(row=1, column=j + 1, value=header)

                    # 写入内容
                    self._write_data_to_worksheet(worksheet, self.time_row_text, time_list)
                    self._write_data_to_worksheet(worksheet, self.name_row_text, name_list)
                    self._write_data_to_worksheet(worksheet, self.uid_row_text, uid_list)
                    self._write_data_to_worksheet(worksheet, self.cont_row_text, cont_list)

                # 删除openpyxl创建的默认工作表
                default_sheet = current_workbook['Sheet']
                current_workbook.remove(default_sheet)

                # 保存工作簿
                if self.out_type == 0:
                    current_workbook.save(os.path.join(self.file_path, get_custom_filename_with_params(
                        self.rule_string, list_name, '', current_count + 1) + f'.xlsx'))
                else:
                    current_workbook.save(os.path.join(self.file_path, "".join(
                        [list_name, '/',
                         get_custom_filename_with_params(self.rule_string, list_name, '', current_count + 1),
                         f'.xlsx'])))
        # 单工作表模式
        else:
            # 设置默认自定义文件名规则
            if len(self.rule_string.strip()) <= 0:
                self.rule_string = '[消息分组]_[消息对象]'
            if '[序号]' in self.rule_string:
                raise ValueError("在单工作表模式中，您无法使用文件命名规则中的[序号]")
            for list_name, object_name_list in files_path.items():
                for i in tqdm(range(len(object_name_list)), f'{list_name} 消息分组导出中'):
                    time_list, name_list, uid_list, cont_list = object_list[object_name_list[i][1]]
                    # 创建workbook对象
                    excel_workbook = openpyxl.Workbook()
                    worksheet = excel_workbook.active
                    worksheet.title = get_custom_filename_with_params(self.sheet_name, list_name,
                                                                      object_name_list[i][0], 0)  # 设置工作表的名字

                    # 写入表头
                    for j, header in enumerate(self.row):
                        worksheet.cell(1, j + 1, header)

                    # 写入内容
                    def write_data_to_worksheet(row_text, data_list):
                        if row_text in self.row:
                            for k, data in enumerate(data_list):
                                worksheet.cell(k + 2, self.row.index(row_text) + 1, data)

                    write_data_to_worksheet(self.time_row_text, time_list)
                    write_data_to_worksheet(self.name_row_text, name_list)
                    write_data_to_worksheet(self.uid_row_text, uid_list)
                    write_data_to_worksheet(self.cont_row_text, cont_list)

                    # 保存工作簿
                    if self.out_type == 0:
                        excel_workbook.save(os.path.join(self.file_path, get_custom_filename_with_params(
                            self.rule_string, list_name, object_name_list[i][0], 0) + f'.xlsx'))
                    else:
                        excel_workbook.save(os.path.join(self.file_path, "".join(
                            [list_name, '/', get_custom_filename_with_params(self.rule_string, list_name,
                                                                             object_name_list[i][0], 0),
                             f'.xlsx'])))

    def toExcel(self) -> None:
        object_file_name_list, object_list = self.get_QQChat_record()
        self._create_excel_file(object_file_name_list, object_list)
