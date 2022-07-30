import os
import re

import openpyxl
from PyQt5.QtCore import QThread, pyqtSignal


def data_clean(text):
    ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')
    text = ILLEGAL_CHARACTERS_RE.sub(r'', text)
    return text


# 清洗windows文件名中的非法字符，只保留中英文和数字
def data_win_file(text):
    ILLEGAL_CHARACTERS_RE = re.compile(r'[^\u4e00-\u9fa5^a-z^A-Z^\d]')
    txt = ILLEGAL_CHARACTERS_RE.sub(r'()', text)
    return txt


def get_QQChat_record(file_path, time_list_out=True, name_list_out=True, uid_list_out=True, cont_list_out=True):
    object_file_name_list = []  # 文件名称列表（分组_备注）
    object_list = []  # 消息对象列表

    time_list = []  # 消息时间列表
    name_list = []  # 昵称列表
    uid_list = []  # QQ or 邮箱列表
    cont_list = []  # 内容列表

    f = open(file_path, encoding="utf-8")
    strs = f.read()
    f.close()

    q_pattern = r'(={64}([\s\S\消息分组:\s\S]{9,32})={64}([\s\S\消息分组:\s\S]{9,32})={64})'  # 定义分隔符
    result = re.split(q_pattern, strs)  # 以pattern的值 分割字符串

    # # 验证是否是4位一循环
    # print(result[0])  # 默认无关内容
    # print(result[1+(4*n)])  # 分组-昵称
    # print(result[2+(4*n)])  # 分组
    # print(result[3+(4*n)])  # 昵称
    # print(result[4+(4*n)])  # 消息内容
    # print((len(result)-1)/4)  # 总人数

    # 多对象获取消息内容
    for i in range(int((len(result) - 1) / 4)):
        object_file_name_list.append(data_win_file(data_clean(
            result[2 + (4 * i)].replace('\n', '').replace('\r', '').replace(' ', '').replace('消息分组:',
                                                                                             ''))) + "_" + data_win_file(
            data_clean(
                result[3 + (4 * i)].replace('\n', '').replace('\r', '').replace(' ', '').replace('消息对象:', ''))))
        # 群消息匹配规则
        pattern = re.compile(
            '(\d{4}-\d{1,2}-\d{1,2}\s\d{1,2}:\d{1,2}:\d{1,2})(.+[(|<](.*)[)|>])([\s\S]*?)(\n\s*\n)')
        # 好友消息匹配规则
        pattern2 = re.compile(
            '(\d{4}-\d{1,2}-\d{1,2}\s\d{1,2}:\d{1,2}:\d{1,2})(.+)([\s\S]*?)(\n\s*\n)')

        # 添加该消息对象各项消息
        m = pattern.findall(result[4 + (4 * i)])
        if len(m) > 0:
            for j in m:
                if time_list_out:
                    time_list.append(j[0])
                if name_list_out:
                    name_list.append(data_clean(j[1].replace(j[2], '').replace('()', '').replace('<>', '')))
                if uid_list_out:
                    uid_list.append(j[2])
                if cont_list_out:
                    cont_list.append(data_clean(j[3][1:]))
        else:
            m = pattern2.findall(result[4 + (4 * i)])
            if len(m) >= 0:
                for j in m:
                    if time_list_out:
                        time_list.append(j[0])
                    if name_list_out:
                        name_list.append(data_clean(j[1].replace(j[2], '').replace('()', '').replace('<>', '')))
                    if uid_list_out:
                        uid_list.append('')
                    if cont_list_out:
                        cont_list.append(data_clean(j[2][1:]))
            else:
                break

        object_list.append([time_list, name_list, uid_list, cont_list])

        # # 输出
        # print(len(time_list),len(name_list),len(uid_list),len(cont_list))
        # for i in range(len(time_list)):
        #     print("time:"+time_list[i]+"\nname:"+name_list[i]+"\nuid:"+uid_list[i]+"\ncont:"+cont_list[i]+"\n===========")
        # print("共：", str(len(time_list)), "条消息")

        # 清空列表
        time_list = []
        name_list = []
        uid_list = []
        cont_list = []

    # print(len(object_file_name_list),len(object_list))  # 判断是否一对象一文件
    return object_file_name_list, object_list


class WorkThread(QThread):
    ProgressRateSignal = pyqtSignal(int)

    def __init__(self, input_file_path: str, out_dir: str, sheet_name: str, out_param: tuple,
                 title_param: tuple):
        """

        :param input_file_path: 输入文件夹地址
        :param out_dir: 输出文件夹地址
        :param sheet_name: 表名
        :param out_param: 输出参数 元组 （时间，昵称，QQ（邮箱），内容）
        :param title_parm: 自定义标题参数元组 （时间，昵称，QQ（邮箱），内容）
        """
        super(WorkThread, self).__init__()
        self.input_file_path = input_file_path
        self.out_dir = out_dir
        self.sheet_name = sheet_name
        self.out_param = out_param
        self.title_param = title_param

    def run(self) -> None:
        """
            QQtoExcel程序
            :param title_param: 自定义标题参数 元组 （时间，昵称，QQ（邮箱），内容）
            :param input_file_path: 输入文件的路径
            :param out_dirs: 输出文件夹的目录
            :param sheet_name: 表名
            :param time_list_out: 是否输出时间
            :param name_list_out: 是否输出姓名
            :param uid_list_out: 是否输出vid
            :param cont_list_out: 是否输出内容
            :return: None
            """
        object_file_name_list, object_list = get_QQChat_record(self.input_file_path,self.out_param[0],
                                                               self.out_param[1], self.out_param[2], self.out_param[3])

        file_path = []  # 输出目录列表
        for i in range(len(object_file_name_list)):
            file_path.append(os.path.join(self.out_dir, object_file_name_list[i] + '.xls'))

        # 写入Excel标题
        row0 = []  # 标题列表
        row1 = []  # 标题序号列表
        if self.out_param[0]:
            row0.append(self.title_param[0])
            row1.append("time_list")
        if self.out_param[1]:
            row0.append(self.title_param[1])
            row1.append("name_list")
        if self.out_param[2]:
            row0.append(self.title_param[2])
            row1.append("uid_list")
        if self.out_param[3]:
            row0.append(self.title_param[3])
            row1.append("cont_list")

        for i in range(len(file_path)):
            time_list = object_list[i][0]
            name_list = object_list[i][1]
            uid_list = object_list[i][2]
            cont_list = object_list[i][3]

            # 创建workbook和sheet对象
            workboot = openpyxl.Workbook()
            worksheet = workboot.active
            worksheet.title = self.sheet_name  # 设置工作表的名字

            for j in range(len(row0)):
                worksheet.cell(1, j + 1, row0[j])

            # 写入内容
            if self.out_param[0]:
                for k in range(len(time_list)):
                    worksheet.cell(k + 2, row1.index("time_list") + 1, time_list[k])
            if self.out_param[1]:
                for k in range(len(name_list)):
                    worksheet.cell(k + 2, row1.index("name_list") + 1, name_list[k])
            if self.out_param[2]:
                for k in range(len(uid_list)):
                    worksheet.cell(k + 2, row1.index("uid_list") + 1, uid_list[k])
            if self.out_param[3]:
                for k in range(len(cont_list)):
                    worksheet.cell(k + 2, row1.index("cont_list") + 1, cont_list[k])

            workboot.save(file_path[i])
            workboot.close()
            self.ProgressRateSignal.emit(int((i+1)/len(file_path)*100))
            if int(i/len(file_path)*100)==97:
                print(1)
